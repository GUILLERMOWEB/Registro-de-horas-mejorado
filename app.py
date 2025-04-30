from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import os
# Carga de variables de entorno desde .env
from dotenv import load_dotenv
load_dotenv()
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from flask_migrate import Migrate
from flask_wtf import FlaskForm
from wtforms import StringField
from wtforms.validators import DataRequired
from wtforms import StringField, SubmitField
from flask_login import login_required, current_user
from functools import wraps

# Importar db de forma tardía para evitar importación circular
from models import db, RegistroHoras, ClienteModel

# Función para convertir una hora en formato de texto a un número decimal
def convertir_hora_a_decimal(hora_str):
    try:
        return float(int(hora_str.strip()))
    except ValueError:
        return 0.0

# Decorador para asegurarse de que solo el superadministrador pueda acceder
def superadmin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if current_user.role != 'superadmin':  # Verifica el rol del usuario
            flash('No tienes permisos para realizar esta acción', 'danger')
            return redirect(url_for('index'))  # Redirige a la página principal
        return f(*args, **kwargs)
    return wrapper

# Inicialización de la aplicación Flask
app = Flask(__name__)

# Configuración de la base de datos con PostgreSQL
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'clave_secreta_para_sesiones'

# Habilita la recarga automática de plantillas y la caché de Jinja
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.cache = {}

# Inicializa la base de datos y el sistema de migración
db.init_app(app)  # Se inicializa db antes de usarlo
migrate = Migrate(app, db)

# Asegúrate de que la base de datos se cree si no existe
with app.app_context():
    db.create_all()

# ─── Modelos ─────────────────────────────────────
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(50), nullable=False)

    registros = db.relationship('Registro', backref='user', lazy=True)


class CentroCosto(db.Model):
    __tablename__ = 'centros_costo'
    id     = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)

class TipoServicio(db.Model):
    __tablename__ = 'tipos_servicio'
    id     = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)

class Linea(db.Model):
    __tablename__ = 'lineas'
    id     = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)

class Registro(db.Model):
    __tablename__ = 'registros'
    __table_args__ = {'extend_existing': True}  # Agrega esta línea

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    fecha = db.Column(db.String(50))
    entrada = db.Column(db.String(50))
    salida = db.Column(db.String(50))
    almuerzo = db.Column(db.Float)
    viaje_ida = db.Column(db.Float, default=0)
    viaje_vuelta = db.Column(db.Float, default=0)
    km_ida = db.Column(db.Float, default=0)
    km_vuelta = db.Column(db.Float, default=0)
    horas = db.Column(db.Float)
    tarea = db.Column(db.Text)
    cliente = db.Column(db.Text)
    comentarios = db.Column(db.Text)
    contrato = db.Column(db.Boolean, default=False)
    centro_costo_id     = db.Column(db.Integer, db.ForeignKey('centros_costo.id'), nullable=True)
    service_order       = db.Column(db.String(10), nullable=True)
    tipo_servicio_id    = db.Column(db.Integer, db.ForeignKey('tipos_servicio.id'), nullable=True)
    linea_id            = db.Column(db.Integer, db.ForeignKey('lineas.id'), nullable=True)
    centro_costo   = db.relationship('CentroCosto')
    tipo_servicio  = db.relationship('TipoServicio')
    linea          = db.relationship('Linea')

    
    
class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    direccion = db.Column(db.String(200), nullable=False)
    telefono = db.Column(db.String(50), nullable=True)

    def __repr__(self):
        return f'<Cliente {self.nombre}>'

        
class FormularioCliente(FlaskForm):
    nombre = StringField('Nombre', validators=[DataRequired()])
    direccion = StringField('Dirección', validators=[DataRequired()])
    telefono = StringField('Teléfono')
    submit = SubmitField('Agregar Cliente')

# ─── Inicialización de la base de datos ─────────
with app.app_context():
    db.create_all()
    if not User.query.filter(db.func.lower(User.username) == 'guillermo gutierrez').first():
        superadmin = User(username='guillermo gutierrez', password='0000', role='superadmin')
        db.session.add(superadmin)
        db.session.commit()

# ─── Rutas ──────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
def inicio():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']

        user = User.query.filter(
            db.func.lower(User.username) == username,
            User.password == password
        ).first()

        if user:
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', category='danger')
    return render_template('login.html')

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        fecha = request.form['fecha']
        entrada = request.form['entrada']
        salida = request.form['salida']

        try:
            almuerzo_horas = int(float(request.form.get('almuerzo_horas', 0)))
            almuerzo = timedelta(hours=almuerzo_horas)
        except ValueError:
            flash("El tiempo de almuerzo debe ser un número válido", "danger")
            return redirect(url_for('dashboard'))

        try:
            viaje_ida = float(request.form.get('viaje_ida', 0) or 0)
            viaje_vuelta = float(request.form.get('viaje_vuelta', 0) or 0)
            km_ida = float(request.form.get('km_ida', 0) or 0)
            km_vuelta = float(request.form.get('km_vuelta', 0) or 0)
        except ValueError:
            flash("Las horas de viaje y kilómetros deben ser números válidos.", "danger")
            return redirect(url_for('dashboard'))

        tarea = request.form.get('tarea', '').strip()
        cliente_nombre = request.form.get('cliente', '').strip()
        comentarios = request.form.get('comentarios', '').strip()

        # Buscar al cliente por nombre
        cliente = ClienteModel.query.filter_by(nombre=cliente_nombre).first()

        if not cliente:
            flash(f"Cliente '{cliente_nombre}' no encontrado.", "danger")
            return redirect(url_for('dashboard'))

        try:
            formato_hora = "%H:%M"
            t_entrada = datetime.strptime(entrada, formato_hora)
            t_salida = datetime.strptime(salida, formato_hora)

            if t_salida < t_entrada:
                t_salida += timedelta(days=1)

            tiempo_total = t_salida - t_entrada - almuerzo
            horas_trabajadas = tiempo_total.total_seconds() / 3600
        except ValueError:
            flash("Formato de hora incorrecto. Use HH:MM.", "danger")
            return redirect(url_for('dashboard'))

        nuevo_registro = Registro(
            user_id=session['user_id'],
            fecha=fecha,
            entrada=entrada,
            salida=salida,
            almuerzo=round(almuerzo.total_seconds() / 3600, 2),
            horas=round(horas_trabajadas, 2),
            viaje_ida=viaje_ida,
            viaje_vuelta=viaje_vuelta,
            km_ida=km_ida,
            km_vuelta=km_vuelta,
            tarea=tarea,
            cliente_id=cliente.id,  # Referencia al cliente por ID
            comentarios=comentarios
        )

        db.session.add(nuevo_registro)
        db.session.commit()
        flash('Registro guardado exitosamente', category='success')
        return redirect(url_for('dashboard'))

    # GET - mostrar los registros y total de horas
    filtros = request.args
    registros_query = Registro.query.filter_by(user_id=session['user_id'])

    if 'fecha' in filtros:
        registros_query = registros_query.filter_by(fecha=filtros['fecha'])

    registros = registros_query.order_by(Registro.fecha.desc()).all()

    total_horas = sum([ 
        (r.horas or 0) + (r.viaje_ida or 0) + (r.viaje_vuelta or 0) 
        for r in registros
    ])
    total_km = sum([
        (r.km_ida or 0) + (r.km_vuelta or 0) 
        for r in registros
    ])

    # Consultar los clientes, centros de costo, tipos de servicio y líneas
    clientes = ClienteModel.query.order_by(ClienteModel.nombre).all()
    centros_costo = CentroCosto.query.order_by(CentroCosto.nombre).all()
    tipos_servicio = TipoServicio.query.order_by(TipoServicio.nombre).all()
    lineas = Linea.query.order_by(Linea.nombre).all()

    return render_template(
        'dashboard.html',
        username=session['username'],
        role=session['role'],
        registros=registros,
        total_horas=round(total_horas, 2),
        total_km=round(total_km, 2),
        clientes=clientes,
        centros_costo=centros_costo,
        tipos_servicio=tipos_servicio,
        lineas=lineas,
        registro=None  # Pasar `None` para crear un nuevo registro
    )





@app.route('/exportar_excel')
def exportar_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    role = session.get('role')

    # 🔵 Obtener filtros de fecha
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    # 🔵 Buscar registros
    if role in ['admin', 'superadmin']:
        registros_query = Registro.query
    else:
        registros_query = Registro.query.filter_by(user_id=session['user_id'])

    if fecha_desde:
        registros_query = registros_query.filter(Registro.fecha >= fecha_desde)
    if fecha_hasta:
        registros_query = registros_query.filter(Registro.fecha <= fecha_hasta)

    registros = registros_query.all()

    # 🔵 Crear DataFrame
    df = pd.DataFrame([{
        'usuario': r.user.username if r.user else 'Usuario eliminado',
        'fecha': r.fecha,
        'entrada': r.entrada,
        'salida': r.salida,
        'almuerzo': r.almuerzo,
        'viaje_ida': r.viaje_ida,
        'viaje_vuelta': r.viaje_vuelta,
        'horas_laborales': r.horas,
        'horas_totales': round((r.horas or 0) + (r.viaje_ida or 0) + (r.viaje_vuelta or 0), 2),
        'km_ida': r.km_ida,
        'km_vuelta': r.km_vuelta,
        'km_totales': (r.km_ida or 0) + (r.km_vuelta or 0),
        'tarea': r.tarea,
        'cliente': r.cliente,
        'comentarios': r.comentarios
    } for r in registros])

    # 🔵 Generar el Excel
    archivo = BytesIO()
    with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Registros')
        ws = writer.sheets['Registros']

        # ✅ Filtros automáticos
        ws.auto_filter.ref = ws.dimensions

        # ✅ Ajuste automático del ancho de columnas
        for col_num, column_cells in enumerate(ws.columns, 1): 
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = adjusted_width

    archivo.seek(0)
    return send_file(archivo, as_attachment=True, download_name=f"registros_{session['username']}.xlsx")




@app.route('/editar_registro/<int:id>', methods=['GET', 'POST'])
def editar_registro(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    registro = Registro.query.get_or_404(id)

    # Cargar listas para los <select>
    centros = CentroCosto.query.order_by(CentroCosto.nombre).all()
    tipos_servicio = TipoServicio.query.order_by(TipoServicio.nombre).all()
    lineas = Linea.query.order_by(Linea.nombre).all()

    if request.method == 'POST':
        # Campos básicos
        fecha = request.form['fecha']
        entrada = request.form['entrada']
        salida = request.form['salida']

        # Viaje y kilómetros
        try:
            viaje_ida = float(request.form.get('viaje_ida', 0) or 0)
            viaje_vuelta = float(request.form.get('viaje_vuelta', 0) or 0)
            km_ida = float(request.form.get('km_ida', 0) or 0)
            km_vuelta = float(request.form.get('km_vuelta', 0) or 0)
        except ValueError:
            flash("Las horas de viaje y kilómetros deben ser números.", "danger")
            return redirect(url_for('editar_registro', id=id))

        # Tarea, cliente y comentarios
        tarea = request.form.get('tarea', '')
        cliente = request.form.get('cliente', '')
        comentarios = request.form.get('comentarios', '')

        # Contrato (Sí/No)
        contrato = bool(int(request.form.get('contrato', 0)))

        # Centro de Costo Contrato
        cc_id = request.form.get('centro_costo_id')
        centro_costo_id = int(cc_id) if cc_id else None

        # Service Order
        service_order = request.form.get('service_order') or None

        # Tipo de Servicio
        ts_id = request.form.get('tipo_servicio_id')
        tipo_servicio_id = int(ts_id) if ts_id else None

        # Línea
        l_id = request.form.get('linea_id')
        linea_id = int(l_id) if l_id else None

        # Almuerzo (horas enteras)
        almuerzo_horas = int(float(request.form.get('almuerzo_horas', 0) or 0))
        almuerzo = timedelta(hours=almuerzo_horas)

        # Cálculo de horas trabajadas
        try:
            t_entrada = datetime.strptime(entrada, "%H:%M")
            t_salida = datetime.strptime(salida, "%H:%M")
            duracion = t_salida - t_entrada - almuerzo
            horas_trabajadas = duracion.total_seconds() / 3600
        except ValueError:
            flash("Error en el formato de hora. Use HH:MM", "danger")
            return redirect(url_for('editar_registro', id=id))

        # Guardar cambios en el registro
        registro.fecha = fecha
        registro.entrada = entrada
        registro.salida = salida
        registro.almuerzo = almuerzo_horas
        registro.horas = round(horas_trabajadas, 2)
        registro.viaje_ida = viaje_ida
        registro.viaje_vuelta = viaje_vuelta
        registro.km_ida = km_ida
        registro.km_vuelta = km_vuelta
        registro.tarea = tarea
        registro.cliente = cliente
        registro.comentarios = comentarios
        registro.contrato = contrato
        registro.centro_costo_id = centro_costo_id
        registro.service_order = service_order
        registro.tipo_servicio_id = tipo_servicio_id
        registro.linea_id = linea_id

        db.session.commit()
        flash('Registro actualizado exitosamente', 'success')

        # Redirigir según rol
        return redirect(url_for('admin') if session['role'] in ['admin', 'superadmin'] else url_for('dashboard'))

    # GET: mostrar formulario con datos y listas
    return render_template('editar_registro.html',
                           registro=registro,
                           centros=centros,
                           tipos_servicio=tipos_servicio,
                           lineas=lineas)




@app.route('/borrar_registro/<int:id>', methods=['POST'])
def borrar_registro(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    registro = Registro.query.get_or_404(id)
    db.session.delete(registro)
    db.session.commit()
    return redirect(url_for('admin') if session['role'] == 'superadmin' else url_for('dashboard'))

@app.route('/crear_admin', methods=['GET', 'POST'])
def crear_admin():
    if 'user_id' not in session or session['role'] != 'superadmin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password')
        confirmar = request.form.get('confirmar_password')

        if not username or not password or not confirmar:
            flash('Todos los campos son obligatorios.', category='warning')
            return render_template('crear_admin.html')

        if password != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('crear_admin.html')

        if User.query.filter_by(username=username).first():
            flash('Ese nombre de usuario ya existe.', category='danger')
        else:
            nuevo_admin = User(username=username, password=password, role='admin')
            db.session.add(nuevo_admin)
            db.session.commit()
            flash('Administrador creado correctamente', category='success')

    return render_template('crear_admin.html')


@app.route('/administrator', methods=['GET', 'POST'])
def admin():
    if 'user_id' not in session or session['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login'))

    filtro_usuario = request.form.get('filtro_usuario') if request.method == 'POST' else None

    usuarios = User.query.with_entities(User.id, User.username).all()

    if filtro_usuario:
        registros = db.session.query(Registro, User).join(User).filter(User.id == filtro_usuario).order_by(Registro.fecha.desc()).all()
    else:
        registros = db.session.query(Registro, User).join(User).order_by(Registro.fecha.desc()).all()

    return render_template('admin.html', registros=registros, usuarios=usuarios,
                           filtro_usuario=filtro_usuario,
                           username=session['username'], role=session['role'])


@app.route('/cambiar_password', methods=['GET', 'POST'])
def cambiar_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nueva = request.form['nueva']
        confirmar = request.form['confirmar']  # Se agrega para la comparación de contraseñas

        if nueva != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('cambiar_password.html')

        # Si las contraseñas coinciden, actualizarla en la base de datos
        user = User.query.get(session['user_id'])
        user.password = nueva
        db.session.commit()
        flash('Contraseña actualizada', category='success')

    return render_template('cambiar_password.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/crear_usuario', methods=['GET', 'POST'])
def crear_usuario():
    if 'user_id' not in session or session['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login'))

    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        confirmar = request.form['confirmar_password']

        if password != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('crear_usuario.html')

        if User.query.filter_by(username=username).first():
            flash('Ese nombre de usuario ya existe.', category='danger')
        else:
            nuevo_usuario = User(username=username, password=password, role='usuario')
            db.session.add(nuevo_usuario)
            db.session.commit()
            flash('Usuario creado exitosamente.', category='success')


    return render_template('crear_usuario.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        password = request.form['password']
        confirmar = request.form['confirmar_password']

        if password != confirmar:
            flash('Las contraseñas no coinciden.', category='danger')
            return render_template('registro.html')

        if User.query.filter_by(username=username).first():
            flash('Ese nombre de usuario ya existe.' , category='danger')
        else:
            nuevo_usuario = User(username=username, password=password, role='usuario')
            db.session.add(nuevo_usuario)
            db.session.commit()
            flash('Usuario creado exitosamente. Ahora podés iniciar sesión.', category='success')
            return redirect(url_for('login'))

    return render_template('registro.html')
    
@app.route('/nuevo_registro', methods=['GET', 'POST'])
def nuevo_registro():
    # Verificar sesión
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Cargar listas para selects
    centros        = CentroCosto.query.order_by(CentroCosto.nombre).all()
    tipos_servicio = TipoServicio.query.order_by(TipoServicio.nombre).all()
    lineas         = Linea.query.order_by(Linea.nombre).all()
    clientes       = ClienteModel.query.order_by(ClienteModel.nombre).all()

    if request.method == 'POST':
        # Campos básicos
        fecha   = request.form['fecha']
        entrada = request.form['entrada']
        salida  = request.form['salida']

        # Viaje y kilómetros
        try:
            viaje_ida    = float(request.form.get('viaje_ida', 0) or 0)
            viaje_vuelta = float(request.form.get('viaje_vuelta', 0) or 0)
            km_ida       = float(request.form.get('km_ida', 0) or 0)
            km_vuelta    = float(request.form.get('km_vuelta', 0) or 0)
        except ValueError:
            flash("Las horas de viaje y kilómetros deben ser números.", "danger")
            return redirect(url_for('nuevo_registro'))

        # Tarea, cliente y comentarios
        tarea       = request.form.get('tarea', '')
        cliente     = request.form.get('cliente', '')
        comentarios = request.form.get('comentarios', '')

        # Campos nuevos
        contrato            = bool(int(request.form.get('contrato', 0)))
        cc_id               = request.form.get('centro_costo')  # Obtener el valor
        centro_costo_id     = int(cc_id) if cc_id else None    # Convertir si existe, sino None
        service_order       = request.form.get('service_order') or None
        ts_id               = request.form.get('tipo_servicio')  # Obtener el valor
        tipo_servicio_id    = int(ts_id) if ts_id else None      # Convertir si existe, sino None
        l_id                = request.form.get('linea')  # Obtener el valor
        linea_id            = int(l_id) if l_id else None  # Convertir si existe, sino None

        # Almuerzo en horas (entero o decimal)
        almuerzo_horas = float(request.form.get('almuerzo_horas', 0) or 0)
        almuerzo = timedelta(hours=almuerzo_horas)

        # Cálculo de horas trabajadas
        try:
            t_entrada = datetime.strptime(entrada, "%H:%M")
            t_salida  = datetime.strptime(salida, "%H:%M")
            duracion  = t_salida - t_entrada - almuerzo
            horas_trabajadas = duracion.total_seconds() / 3600
        except ValueError:
            flash("Error en el formato de hora. Use HH:MM", "danger")
            return redirect(url_for('nuevo_registro'))

        # Crear y guardar registro
        nuevo_registro = Registro(
            user_id=session['user_id'],
            fecha=fecha,
            entrada=entrada,
            salida=salida,
            almuerzo=almuerzo_horas,
            horas=round(horas_trabajadas, 2),
            viaje_ida=viaje_ida,
            viaje_vuelta=viaje_vuelta,
            km_ida=km_ida,
            km_vuelta=km_vuelta,
            tarea=tarea,
            cliente=cliente,
            comentarios=comentarios,
            contrato=contrato,
            centro_costo_id=centro_costo_id,  # Aquí se integra
            service_order=service_order,
            tipo_servicio_id=tipo_servicio_id,  # Aquí se integra
            linea_id=linea_id  # Aquí se integra
        )

        db.session.add(nuevo_registro)
        db.session.commit()
        flash('Registro creado exitosamente', 'success')

        # Redirigir según rol
        return redirect(url_for('admin') if session.get('role') in ['admin', 'superadmin'] else url_for('dashboard'))

    # GET - mostrar formulario vacío
    return render_template(
        'nuevo_registro.html',
        centros=centros,
        tipos_servicio=tipos_servicio,
        lineas=lineas,
        clientes=clientes
    )

@app.route('/usuarios')
def listar_usuarios():
    if 'user_id' not in session or session['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login'))

    usuarios = User.query.with_entities(User.id, User.username, User.role).all()
    return render_template('usuarios.html', usuarios=usuarios)
    
# ─── CRUD Clientes (solo superadmin) ───────────────────────────

@@app.route('/ver_cliente', methods=['GET', 'POST'])
def ver_cliente():
    clientes = Cliente.query.all()  # Obtener todos los clientes

    if request.method == 'POST':
        cliente_id = request.form['cliente']  # Obtener el ID del cliente seleccionado
        
        if not cliente_id:
            flash('Debe seleccionar un cliente.', 'danger')
            return redirect(url_for('ver_cliente'))
        
        cliente = Cliente.query.get(cliente_id)  # Obtener el cliente por su ID

        if cliente:
            return render_template('detalle_cliente.html', cliente=cliente)  # Muestra los detalles del cliente
        else:
            flash('Cliente no encontrado.', 'danger')
            return redirect(url_for('ver_cliente'))  # Redirige de vuelta si no se encuentra el cliente

    return render_template('ver_cliente.html', clientes=clientes)


@app.route('/agregar_cliente', methods=['GET', 'POST'])
@login_required
def agregar_cliente():
    # Solo el superadmin puede agregar clientes
    if current_user.role != 'superadmin':
        flash('Acceso denegado: solo el superadministrador puede agregar clientes.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        nombre = request.form['nombre']
        direccion = request.form['direccion']
        telefono = request.form.get('telefono')

        # Aquí cambiamos ClienteModel a Cliente
        nuevo_cliente = Cliente(nombre=nombre, direccion=direccion, telefono=telefono)

        try:
            db.session.add(nuevo_cliente)
            db.session.commit()
            flash('Cliente agregado exitosamente.', 'success')
            return redirect(url_for('dashboard'))  # <-- Redirige al dashboard superadmin
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar el cliente: {e}', 'danger')

    # Traer todos los clientes para mostrar en el formulario si querés
    # Aquí también cambiamos ClienteModel a Cliente
    clientes = Cliente.query.all()

    return render_template('agregar_cliente.html', clientes=clientes)


    

@app.route('/editar_cliente/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    # Verificación de rol
    if session.get('role') != 'superadmin':
        flash("Acceso denegado: solo el superadministrador puede editar clientes.", "danger")
        return redirect(url_for('dashboard'))

    cliente = ClienteModel.query.get_or_404(cliente_id)

    if request.method == 'POST':
        nuevo_nombre = request.form.get('nombre', '').strip()
        nueva_direccion = request.form.get('direccion', '').strip()
        nuevo_telefono = request.form.get('telefono', '').strip()

        # Validaciones básicas
        if not nuevo_nombre:
            flash("El nombre no puede estar vacío.", "danger")
        elif ClienteModel.query.filter(
            ClienteModel.nombre == nuevo_nombre,
            ClienteModel.id != cliente_id
        ).first():
            flash("Ya existe otro cliente con ese nombre.", "warning")
        else:
            cliente.nombre = nuevo_nombre
            cliente.direccion = nueva_direccion
            cliente.telefono = nuevo_telefono

            try:
                db.session.commit()
                flash("Cliente actualizado con éxito.", "success")
                return redirect(url_for('ver_cliente'))
            except Exception as e:
                db.session.rollback()
                flash(f"Error al actualizar el cliente: {e}", "danger")

    return render_template('editar_cliente.html', cliente=cliente)

@app.route('/borrar_cliente/<int:cliente_id>', methods=['POST'])
@login_required
def borrar_cliente(cliente_id):
    # Solo permite si el usuario es superadmin
    if session.get('role') != 'superadmin':
        flash("Acceso denegado: solo el superadministrador puede eliminar clientes.", "danger")
        return redirect(url_for('dashboard'))

    try:
        cliente = ClienteModel.query.get_or_404(cliente_id)
        db.session.delete(cliente)
        db.session.commit()
        flash("Cliente eliminado correctamente.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar el cliente: {e}", "danger")

    return redirect(url_for('ver_cliente'))


with app.app_context():
    db.create_all()



if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
